import openpyxl
from pythonutils.hidden_utils import *
from pythonutils.os_utils import *
from pythonutils.yes_or_no_input import *

_verbose = False


def get_column_index_of_header_with_value(sheet, header_value):
    if _verbose:
        print('get_column_index_of_header_with_value ( sheet: {} , header_value: {} )'.format(sheet.name, header_value))

    for i in range(1, sheet.max_column + 1):
        cell = sheet.cell(row=1, column=i)
        if cell.value == header_value:
            return i

    return -1


def get_row_index_of_column_with_value(sheet, column_index, target_value):
    if _verbose:
        print('get_row_index_of_column_with_value ( sheet: {} , column_index: {} , target_value: {} )'.format(
            sheet.name, column_index, target_value))

    for i in range(1, sheet.max_row + 1):
        cell = sheet.cell(row=i, column=column_index)
        if cell.value == target_value:
            return i

    return -1


def remove_temporary_workbooks(paths):
    if _verbose:
        print('remove_temporary_workbooks ( paths: {} )'.format(len(paths)))

    result = []

    for path in paths:
        file_name = get_file_name_from_path(path)
        if not file_name.startswith('~$'):
            result.append(path)

    return result


def clone_sheet(from_this_workbook, to_this_workbook, sheet_name):
    if _verbose:
        print('clone_sheet ( from_this_workbook: {} , to_this_workbook: {} , sheet_name: {} )'.format(from_this_workbook.name, to_this_workbook, sheet_name))

    # todo remove this special case
    sheet_title = sheet_name
    if sheet_title == 'Currencies':
        sheet_title = 'Currencies -- Revamp'
    to_this_workbook.create_sheet(title=sheet_title, index=len(to_this_workbook.sheetnames))

    from_this_workbook_sheet = from_this_workbook[sheet_name]
    to_this_workbook_sheet = to_this_workbook[sheet_name]

    max_row = from_this_workbook_sheet.max_row
    max_column = from_this_workbook_sheet.max_column
    row = 0
    for row in range(1, max_row + 1):
        # iterate over all columns
        for column in range(1, max_column + 1):
            from_this_workbook_sheet_cell = from_this_workbook_sheet.cell(row, column).value

            # todo remove this special case
            if type(from_this_workbook_sheet_cell) is str:
                from_this_workbook_sheet_cell = from_this_workbook_sheet_cell.replace('PackName[Pack Name]', 'PackName')

            to_this_workbook_sheet.cell(row, column).value = from_this_workbook_sheet_cell

    return to_this_workbook


def get_all_sheet_names_that_startswith(from_this_workbook, sheet_name):
    if _verbose:
        print('get_all_sheet_names_that_startswith ( sheet_name: {} )'.format(sheet_name))

    results = []

    for sheetname in from_this_workbook.sheetnames:
        if sheetname.startswith(sheet_name):
            results.append(sheetname)

    return results


def find_sheet_name_that_startswith(from_this_workbook, sheet_name):
    if _verbose:
        print('find_sheet_name_that_startswith ( sheet_name: {} )'.format(sheet_name))

    sheet_names_that_startswith = get_all_sheet_names_that_startswith(from_this_workbook, sheet_name)
    sheet_names_that_startswith_len = len(sheet_names_that_startswith)

    if sheet_names_that_startswith_len > 1:
        print('There are multiple sheets that start with {}. Choose which one should be targeted.'.format(sheet_name))

        for sheet_name_that_startswith in sheet_names_that_startswith:
            if yes_or_no('Is {} the sheet that you want?'.format(sheet_name_that_startswith)):
                return sheet_name_that_startswith
    elif sheet_names_that_startswith_len == 1:
        return sheet_names_that_startswith[0]

    print('Could not find any sheet that starts with {}'.format(sheet_name))
    return None


def get_row_index_of_first_empty_cell_in_column(from_this_sheet, in_this_column_index):
    if _verbose:
        print('get_row_index_of_first_empty_cell_in_column ( in_this_column_index: {} )'.format(
            in_this_column_index))

    max_row = from_this_sheet.max_row
    for row in range(1, max_row + 1):
        if from_this_sheet.cell(row, in_this_column_index).value is None:
            return row


def get_column_indexes_of_header_rows_that_startswith(from_this_sheet, starts_with):
    if _verbose:
        print('get_column_indexes_of_header_rows_that_startswith ( starts_with: {} )'.format(
            starts_with))

    indexes = []

    max_column = from_this_sheet.max_column
    for column in range(1, max_column + 1):
        cell_value = from_this_sheet.cell(1, column).value

        if type(cell_value) is unicode:
            cell_value = str(cell_value)

        if type(cell_value) is str:
            if cell_value.startswith(starts_with):
                indexes.append(column)

    return indexes


def find_column_index_of_header_rows_that_startswith(from_this_sheet, starts_with):
    if _verbose:
        print('find_column_index_of_header_rows_that_startswith ( starts_with: {} )'.format(
            starts_with))

    column_indexes_of_header_rows_that_startswith = get_column_indexes_of_header_rows_that_startswith(
        from_this_sheet, starts_with)
    column_indexes_of_header_rows_that_startswith_len = len(column_indexes_of_header_rows_that_startswith)

    if column_indexes_of_header_rows_that_startswith_len > 1:
        print('There are multiple headers that start with {}. Choose which one should be targeted.'.format(starts_with))

        for column_index_of_header_rows_that_startswith in column_indexes_of_header_rows_that_startswith:
            cell_value = from_this_sheet.cell(1, column_index_of_header_rows_that_startswith).value
            if yes_or_no('Is {} the header that you want?'.format(cell_value)):
                return column_index_of_header_rows_that_startswith
    elif column_indexes_of_header_rows_that_startswith_len == 1:
        return column_indexes_of_header_rows_that_startswith[0]

    print('Could not find any header that starts with {}'.format(starts_with))
    return None


def transpose_row(from_sheet, from_row, to_sheet, to_row):
    if _verbose:
        print('transpose_row ( from_row: {} , to_row: {} )'.format(from_row, to_row))

    from_headers_and_values = {}

    max_column = from_sheet.max_column
    for column in range(1, max_column + 1):
        from_headers_and_values[from_sheet.cell(1, column).value] = from_sheet.cell(from_row, column).value

    max_column = to_sheet.max_column
    for column in range(1, max_column + 1):
        to_header = to_sheet.cell(1, column).value
        if to_header in from_headers_and_values:
            print(to_header)
            del from_headers_and_values[to_header]

    print(from_headers_and_values)


def get_broken_named_ranges(named_ranges):
    if _verbose:
        print('get_broken_named_ranges ( named_ranges: {} )'.format(named_ranges))

    broken_named_ranges = []
    for named_range in named_ranges:
        if '#REF' in named_range.attr_text:
            broken_named_ranges.append(named_range)

    return broken_named_ranges


def delete_broken_named_ranges(target_workbook_path):
    if _verbose:
        print('delete_broken_named_ranges ( target_workbook_path: {} )'.format(target_workbook_path))

    workbook = openpyxl.load_workbook(target_workbook_path, data_only=False)
    broken_named_ranges = get_broken_named_ranges(workbook.defined_names.definedName)
    while len(broken_named_ranges) > 0:
        for broken_named_range in broken_named_ranges:
            print(broken_named_range.name)
            print(broken_named_range)
            print('')
            workbook.defined_names.delete(broken_named_range.name, broken_named_range.localSheetId)
        broken_named_ranges = get_broken_named_ranges(workbook.defined_names.definedName)

    workbook.save(target_workbook_path)
    workbook.close()
