import openpyxl
from pythonutils.colors_utils import *
from pythonutils.os_utils import *
from pythonutils.str_utils import *

_verbose = False
_super_verbose = False


def search_dir(path, term, search_formula, case_sensitive, match_entire_val, sheet_name_contains, sheet_name_excludes,
               workbook_name_contains, workbook_name_excludes, recursive, colorize):
    if _verbose:
        print("search_dir ( path: {} , term: {} , search_formula: {} , case_sensitive: {} , match_entire_val: {} , "
              "sheet_name_contains: {} , sheet_name_excludes: {} , workbook_name_contains: {} , "
              "workbook_name_excludes: {} , recursive: {} , colorize: {} )".format(path, term, search_formula,
                                                                                   case_sensitive, match_entire_val,
                                                                                   sheet_name_contains,
                                                                                   sheet_name_excludes,
                                                                                   workbook_name_contains,
                                                                                   workbook_name_excludes, recursive,
                                                                                   colorize))

    search_dir_results = []
    # matching workbook filter will not support case sensitivity
    workbook_name_contains = workbook_name_contains.lower()
    workbook_name_excludes = workbook_name_excludes.lower()
    workbook_paths = []

    # remove workbooks that do not need to be searched
    workbook_paths_to_check = get_all_in_dir(target_dir=path, full_path=True, recursive=recursive, include_dirs=False, include_files=True, must_end_in=".xlsx")
    for workbook_path in workbook_paths_to_check:
        workbook_name = get_file_name_from_path(workbook_path)
        workbook_name_lower = workbook_name.lower()

        remove_workbook = workbook_name.startswith("~") or \
                          (workbook_name_contains != "" and workbook_name_contains not in workbook_name_lower) or \
                          (workbook_name_excludes != "" and workbook_name_excludes in workbook_name_lower)

        if _verbose:
            remove_workbook_debug = get_green(remove_workbook) if remove_workbook is False else \
                get_red(remove_workbook)
            print("\tRemoval of {}: {}".format(workbook_path, remove_workbook_debug))

        if not remove_workbook:
            workbook_paths.append(workbook_path)

    workbook_path_count = len(workbook_paths)
    workbook_paths_searched = 0
    for workbook_path in workbook_paths:
        workbook_paths_searched += 1
        workbook_name = get_file_name_from_path(workbook_path)

        print("{}/{} searching workbook: {}".format(num_to_comma_str(
            workbook_paths_searched), num_to_comma_str(workbook_path_count), get_yellow(workbook_name)))

        search_file_results = search_workbook(workbook_path, term, search_formula, case_sensitive, match_entire_val,
                                              sheet_name_contains, sheet_name_excludes, colorize)
        search_dir_results.extend(search_file_results)

    return search_dir_results


def search_workbook(path, term, search_formula, case_sensitive, match_entire_val, sheet_name_contains,
                    sheet_name_excludes, colorize):
    if _verbose:
        print("search_workbook ( path: {} , term: {} , search_formula: {} , case_sensitive: {} , " \
              "match_entire_val: {} , sheet_name_contains: {} , sheet_name_excludes: {} , colorize: {} )".format(
                path, term,
                search_formula, case_sensitive,
                match_entire_val, sheet_name_contains,
                sheet_name_excludes, colorize))

    if not case_sensitive:
        term = term.lower()

    search_file_results = []
    # data_only set to false will have cell formulas returned from .value
    workbook = openpyxl.load_workbook(path, data_only=False if search_formula else True)
    workbook_name = get_file_name_from_path(path)

    sheet_name_contains = sheet_name_contains.lower()
    sheet_name_excludes = sheet_name_excludes.lower()

    workbook_sheet_count = len(workbook.sheetnames)
    workbook_sheets_searched = 0
    sheet_names = []

    # remove sheets that do not need to be searched
    sheet_names_to_check = workbook.sheetnames
    for sheet_name in sheet_names_to_check:
        sheet_name_lower = sheet_name.lower()
        remove_sheet = (sheet_name_contains != "" and sheet_name_contains not in sheet_name_lower) or \
                       (sheet_name_excludes != "" and sheet_name_excludes in sheet_name_lower)

        if _verbose:
            remove_sheet_debug = get_green(remove_sheet) if remove_sheet is False else \
                get_yellow(remove_sheet)
            print("\tRemoval of {}: {}".format(sheet_name, remove_sheet_debug))

        if not remove_sheet:
            sheet_names.append(sheet_name)

    for sheet_name in sheet_names:
        sheet = workbook.get_sheet_by_name(sheet_name)

        # iterate over all rows
        max_row = sheet.max_row
        max_column = sheet.max_column
        cell_count = max_row * max_column

        workbook_sheets_searched += 1
        print("\t{}/{} searching sheet: {} with {} cells".format(num_to_comma_str(
            workbook_sheets_searched), num_to_comma_str(workbook_sheet_count), get_green(sheet_name),
            num_to_comma_str(cell_count)))

        row = 0
        for row in range(1, max_row + 1):
            # iterate over all columns
            for column in range(1, max_column + 1):
                col_letter = openpyxl.utils.get_column_letter(column)
                cell = sheet.cell(row, column)
                cell_val = cell.value
                str_cell_val = str(cell_val.encode("utf8")) if type(cell_val) is str else str(cell_val)
                str_cell_val_to_compare = str_cell_val
                if not case_sensitive:
                    str_cell_val_to_compare = str_cell_val_to_compare.lower()

                cell_result = "workbook: {} | sheet: {} | cell {}{}: {}".format(
                    (get_yellow(workbook_name) if colorize else workbook_name),
                    (get_green(sheet_name) if colorize else sheet_name),
                    col_letter, row, (get_blue(str_cell_val) if colorize else str_cell_val))

                if _super_verbose:
                    print(cell_result)

                add = term == str_cell_val_to_compare if match_entire_val else term in str_cell_val_to_compare
                if add:
                    search_file_results.append(cell_result)

    workbook.close()

    return search_file_results
